"""Tests for RAID agent and Excel export."""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from agentic_ai.raid_agent import (
    raid_dict_to_markdown,
    run_raid_agent,
    write_raid_excel,
)


@dataclass
class _Msg:
    content: str


@dataclass
class _Choice:
    message: _Msg


@dataclass
class _Resp:
    choices: list[_Choice]


class _FakeCompletions:
    def __init__(self, payload: dict) -> None:
        self._payload = payload

    def create(self, model: str, messages: list, **kwargs: object) -> _Resp:
        return _Resp([_Choice(_Msg(json.dumps(self._payload)))])


class _FakeOpenAI:
    def __init__(self, payload: dict) -> None:
        self.chat = SimpleNamespace(completions=_FakeCompletions(payload))


def test_write_raid_excel_roundtrip(tmp_path: Path) -> None:
    data = {
        "title": "T",
        "risks": [{"id": "R1", "title": "r", "description": "d", "severity": "High", "owner": "o", "mitigation": "m", "due": ""}],
        "assumptions": [],
        "issues": [],
        "dependencies": [],
    }
    p = tmp_path / "r.xlsx"
    write_raid_excel(p, data)
    wb = load_workbook(p)
    ws = wb.active
    assert ws["A1"].value == "Category"
    assert ws["A2"].value == "Risk"


def test_raid_dict_to_markdown() -> None:
    md = raid_dict_to_markdown({"title": "X", "risks": [{"title": "R", "description": "D"}]})
    assert "# X" in md
    assert "## Risks" in md


def test_run_raid_agent_with_fake_llm(tmp_path: Path) -> None:
    payload = {
        "title": "RAID test",
        "risks": [{"id": "R1", "title": "Risk A", "description": "D", "severity": "Medium", "owner": "PM", "mitigation": "M", "due": ""}],
        "assumptions": [{"id": "A1", "title": "Asm", "description": "D", "severity": "Low", "owner": "PM", "mitigation": "M", "due": ""}],
        "issues": [{"id": "I1", "title": "Iss", "description": "D", "severity": "High", "owner": "TL", "mitigation": "M", "due": ""}],
        "dependencies": [{"id": "D1", "title": "Dep", "description": "D", "severity": "Medium", "owner": "Ops", "mitigation": "M", "due": ""}],
    }
    art = run_raid_agent(
        meeting_text="Kickoff notes.",
        scope_text="Scope baseline.",
        output_dir=tmp_path,
        use_llm=True,
        llm_client=_FakeOpenAI(payload),
    )
    assert art.llm_used is True
    assert art.excel_file.is_file()
    assert art.summary_file.is_file()


def test_run_raid_agent_template_without_llm(tmp_path: Path) -> None:
    art = run_raid_agent(
        meeting_text="m",
        scope_text="s",
        output_dir=tmp_path,
        use_llm=False,
    )
    assert art.llm_used is False
    assert art.excel_file.exists()
